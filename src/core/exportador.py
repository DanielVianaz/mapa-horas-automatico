from openpyxl import load_workbook
import os
import sys
import calendar
import datetime
import holidays
from tkinter import messagebox


# =========================
# CONFIG
# =========================
COORDENADOR = "VERA BRAZ"
NOME_ABA = "Planilha1"


# =========================
# CAMINHO BASE (FUNCIONA NO .EXE)
# =========================
def get_caminho_base():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


# =========================
# DATAS E REGRAS
# =========================
def obter_ultimo_dia_mes(ano, mes):
    return calendar.monthrange(ano, mes)[1]


def contar_dias_uteis(ano, mes):
    ultimo_dia = obter_ultimo_dia_mes(ano, mes)
    feriados = holidays.Portugal(years=ano)

    dias_uteis = 0

    for dia in range(1, ultimo_dia + 1):
        data = datetime.date(ano, mes, dia)

        if data.weekday() < 5 and data not in feriados:
            dias_uteis += 1

    return dias_uteis


def calcular_horas_mensais(ano, mes):
    return contar_dias_uteis(ano, mes) * 8


# =========================
# FUNÇÃO PRINCIPAL
# =========================
def exportar_excel(mapa, ano, mes):
    base_path = get_caminho_base()

    caminho_template = os.path.join(
        base_path,
        "data",
        "template",
        "template.xlsx"
    )

    # 🔥 VALIDAÇÃO DO TEMPLATE (evita erro silencioso)
    if not os.path.exists(caminho_template):
        raise FileNotFoundError(f"Template não encontrado em:\n{caminho_template}")

    pasta_saida = os.path.join(
        os.path.expanduser("~"),
        "Desktop",
        "Mapas_Gerados"
    )
    os.makedirs(pasta_saida, exist_ok=True)

    arquivos_gerados = []

    for nome, info in mapa.items():

        wb = load_workbook(caminho_template)
        ws = wb[NOME_ABA] if NOME_ABA in wb.sheetnames else wb.active

        # =========================
        # CABEÇALHO
        # =========================
        preencher_cabecalho(ws, nome, ano, mes)

        # =========================
        # DIAS
        # =========================
        preencher_dias(ws, info)

        # =========================
        # LIMPAR TOTAIS
        # =========================
        limpar_totais(ws)

        # =========================
        # SALVAR
        # =========================
        nome_arquivo = f"{nome.replace(' ', '_')}_{mes}_{ano}.xlsx"
        caminho_saida = os.path.join(pasta_saida, nome_arquivo)

        wb.save(caminho_saida)
        arquivos_gerados.append(caminho_saida)

    messagebox.showinfo(
        "Sucesso",
        f"Mapas gerados com sucesso!\n\nArquivos salvos em:\n{pasta_saida}"
    )

    return arquivos_gerados


# =========================
# CABEÇALHO
# =========================
def preencher_cabecalho(ws, nome, ano, mes):
    ws["B1"] = ano
    ws["B2"] = mes

    dias_uteis = contar_dias_uteis(ano, mes)
    ws["B3"] = dias_uteis

    horas_mensais = calcular_horas_mensais(ano, mes)
    ws["B4"] = horas_mensais

    ws["B6"] = COORDENADOR
    ws["A10"] = nome

    ultimo_dia = obter_ultimo_dia_mes(ano, mes)
    data_formatada = f"{ultimo_dia:02d}/{mes:02d}/{ano}"

    ws["A46"] = data_formatada


# =========================
# DIAS
# =========================
def preencher_dias(ws, info):
    linha_inicio = 12

    for dia_info in info["dias"]:
        linha = linha_inicio + dia_info["dia"] - 1

        ws.cell(row=linha, column=4).value = dia_info["turno"]
        ws.cell(row=linha, column=5).value = dia_info["horas"]
        ws.cell(row=linha, column=6).value = dia_info["noturno"]


# =========================
# LIMPAR TOTAIS
# =========================
def limpar_totais(ws):
    ws["E44"] = ""
    ws["F44"] = ""